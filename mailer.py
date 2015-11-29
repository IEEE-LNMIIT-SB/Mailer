from __future__ import print_function

__author__      = "Vishwajeet Narwal (Macbull)"
__email__ = "vnarwal95@gmail.com"


from apiclient import discovery
from apiclient import errors
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from msg import msg,count,excel,subject,limit,salutation
from oauth2client import client
from oauth2client import tools
import base64
import httplib2
import oauth2client
import openpyxl
import os

try:
    import argparse
    flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
except ImportError:
    flags = None

SCOPES = 'https://www.googleapis.com/auth/gmail.send'
CLIENT_SECRET_FILE = 'client_secret.json'
APPLICATION_NAME = 'Macbull Mailer'

def get_credentials(count):
    home_dir = os.getcwd()
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'gmail'+str(count)+'.json')

    store = oauth2client.file.Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        if flags:
            credentials = tools.run_flow(flow, store, flags)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print('Storing credentials to ' + credential_path)
    return credentials

def CreateMessage(sender, to, subject, message_text):
  message = MIMEText(message_text)
  message['to'] = to
  message['from'] = sender
  message['subject'] = subject
  return {'raw': base64.urlsafe_b64encode(message.as_string())}

def SendMessage(service, user_id, message):
    try:
        message = (service.users().messages().send(userId=user_id, body=message).execute())
        print ('Message Id: %s' % message['id'])
        return message
    except errors.HttpError, error:
        print ('An error occurred: %s' % error)

def DraftMessage(salutation,name,msg):
    msg=salutation+" "+name+"\n"+msg
    return msg

def readExcel(listpath):
    email=[]
    names=[]
    wb = openpyxl.load_workbook(listpath)
    sheetdata=wb.active
    if sheetdata.max_column==1:
        name=False
    else:
        name=True
    for i in range(1,sheetdata.max_row+1,1):
        if name:
            names.append(sheetdata.cell(row=i, column=1).value)
            email.append(sheetdata.cell(row=i, column=2).value)  
        else:
            email.append(sheetdata.cell(row=i, column=1).value)  
        

    return name,names,email
                
def main():
    i=0
    credentials=[]
    http=[]
    service=[]
    quantity={}

    name,names,email=readExcel(excel)
    while i < count:
        credentials.append(get_credentials(i))
        http.append(credentials[i].authorize(httplib2.Http()))
        service.append(discovery.build('gmail', 'v1', http=http[i]))
        quantity[i]=0
        i=i+1
    i=0
    while i < len(email):
        j=0
        old_i=i
        while j < count:
            if quantity[j]<limit:
                if name:
                    draft=DraftMessage(salutation,names[i],msg)
                else:
                    draft=DraftMessage(salutation,"",msg)
                message=CreateMessage("no-reply@plinth.com",email[i],subject,draft)
                mes=SendMessage(service[j],'me',message)
                print('Sent to ',email[i])
                quantity[j]+=1
                i=i+1
            j=j+1
        if i==old_i:
            print("Limit reached for all account")
            print ("Last Email Sent to "+email[i])

if __name__ == '__main__':
    main()